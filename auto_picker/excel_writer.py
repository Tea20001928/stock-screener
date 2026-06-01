"""
自动选股 - Excel 输出（带颜色标记）
绿色=优选，黄色=合格，红色=不合格，灰色=不达标排最后
"""

import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    Alignment,
    Border,
    Side,
    PatternFill,
)
from openpyxl.utils import get_column_letter

from auto_picker.config import (
    OUTPUT_DIR,
    OUTPUT_DIR_NAME,
    COLUMNS,
    COLOR_GREEN,
    COLOR_YELLOW,
    COLOR_RED,
    COLOR_GREY,
)
from data_fetcher import is_using_real_free_float


def _ensure_output_dir():
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)


def write_excel(results: list, sanban_date: str, erban_date: str, yiban_date: str) -> str:
    """
    生成带颜色标记的 Excel 报表

    参数:
        results: 股票数据列表（含内部标记字段）
        sanban_date: 三板日期
        erban_date: 二板日期
        yiban_date: 一板日期

    返回:
        Excel 文件路径
    """
    _ensure_output_dir()

    date_label = sanban_date.replace("-", "")
    filename = f"自动选股_{date_label}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    # 如果文件已存在，加时间戳
    if os.path.exists(filepath):
        timestamp = datetime.now().strftime("%H%M%S")
        filename = f"自动选股_{date_label}_{timestamp}.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "自动选股"

    # 自由流通市值 vs 流通市值 回退
    cap_label = "二板自由流通市值" if is_using_real_free_float() else "二板流通市值"
    col_rename = {}
    if "二板自由流通市值" in COLUMNS and cap_label != "二板自由流通市值":
        col_rename["二板自由流通市值"] = cap_label
    # 反向映射：display name → internal key
    internal_key = {v: k for k, v in col_rename.items()}
    display_columns = [col_rename.get(col, col) for col in COLUMNS]

    # === 样式定义 ===
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="微软雅黑", size=10)
    data_alignment = Alignment(horizontal="center", vertical="center")
    data_alignment_left = Alignment(horizontal="left", vertical="center")
    data_alignment_right = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # 颜色填充
    def make_fill(hex_color):
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    fill_green = make_fill(COLOR_GREEN)
    fill_yellow = make_fill(COLOR_YELLOW)
    fill_red = make_fill(COLOR_RED)
    fill_grey = make_fill(COLOR_GREY)

    # === 标题行 ===
    title = f"自动选股报告 — 三板 {sanban_date} | 二板 {erban_date} | 一板 {yiban_date}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(display_columns))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # 副标题（含颜色图例）
    subtitle = (
        f"共 {len(results)} 只候选 | "
        f"绿色=优选 | 黄色=合格 | 红色=不合格 | 灰色=不达标(三板竞价不满足基础门槛) | "
        f"生成: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(display_columns))
    sub_cell = ws.cell(row=2, column=1, value=subtitle)
    sub_cell.font = Font(name="微软雅黑", size=9, color="808080")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # === 表头（第3行）===
    header_row = 3
    for col_idx, col_name in enumerate(display_columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    ws.row_dimensions[header_row].height = 38

    # === 数据行 ===
    percent_cols = {"三板竞价涨幅", "二板竞价涨幅", "指标2:竞流比(%)"}
    amount_cols = {"三板竞价金额", "二板竞价金额", cap_label}
    ratio_cols = {"指标1:三板/二板竞价金额"}

    for row_idx, stock in enumerate(results):
        excel_row = row_idx + 4
        ws.row_dimensions[excel_row].height = 26

        # 整行颜色
        overall_color = stock.get("_overall_color", COLOR_GREY)
        row_fill = make_fill(overall_color)

        for col_idx, col_name in enumerate(display_columns, start=1):
            lookup_key = internal_key.get(col_name, col_name)
            value = stock.get(lookup_key, "N/A")
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.fill = row_fill

            # 对齐和数字格式
            if col_name in percent_cols or col_name in ratio_cols:
                cell.alignment = data_alignment_right
                if isinstance(value, (int, float)):
                    if "涨幅" in col_name:
                        cell.number_format = '0.00"%"'
                    elif "竞流比" in col_name:
                        cell.number_format = '0.0000"%"'
                    else:
                        cell.number_format = "0.0000"
            elif col_name in amount_cols:
                cell.alignment = data_alignment_right
                if isinstance(value, (int, float)):
                    cell.number_format = "#,##0"
            elif col_name == "股票代码":
                cell.alignment = data_alignment
                cell.number_format = "@"
            elif col_name == "股票名称":
                cell.alignment = data_alignment_left
            elif col_name == "行业":
                cell.alignment = data_alignment_left
            elif col_name == "评级":
                cell.alignment = data_alignment
                # 评级加粗
                cell.font = Font(name="微软雅黑", size=10, bold=True)
            elif col_name == "二板最后涨停时间":
                cell.alignment = data_alignment
            else:
                cell.alignment = data_alignment

    # === 列宽 ===
    col_widths = {
        "股票代码": 12,
        "股票名称": 12,
        "行业涨停数": 10,
        "三板竞价涨幅": 14,
        "三板竞价金额": 16,
        "二板竞价涨幅": 14,
        "二板竞价金额": 16,
        "二板自由流通市值": 16,
        "二板流通市值": 16,
        "指标1:三板/二板竞价金额": 20,
        "指标2:竞流比(%)": 16,
        "二板最后涨停时间": 16,
        "行业": 12,
        "评级": 10,
    }
    for col_idx, col_name in enumerate(display_columns, start=1):
        width = col_widths.get(col_name, 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # === 冻结和筛选 ===
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(display_columns))}{len(results) + header_row}"

    wb.save(filepath)
    return filepath
