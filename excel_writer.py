"""
Excel 输出模块 - 生成格式化的 Excel 报表
"""

import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    Alignment,
    Border,
    Side,
    PatternFill,
    numbers,
)
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR, OUTPUT_DIR_NAME, COLUMNS
from data_fetcher import is_using_real_free_float


def _ensure_output_dir():
    """确保输出文件夹存在"""
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)


def _format_amount(value):
    """
    格式化金额为人民币显示
    大于1亿的显示为 'X.XX 亿'
    大于1万的显示为 'X.XX 万'
    """
    if not isinstance(value, (int, float)):
        return value
    if value >= 1e8:
        return f"{value / 1e8:.2f} 亿"
    elif value >= 1e4:
        return f"{value / 1e4:.2f} 万"
    else:
        return f"{value:.2f}"


def write_excel(stock_list: list, current_date: str, prev_date: str) -> str:
    """
    生成 Excel 报表

    参数:
        stock_list: 股票数据列表
        current_date: 二板日期 (YYYY-MM-DD)
        prev_date: 一板日期 (YYYY-MM-DD)

    返回:
        生成的 Excel 文件路径
    """
    _ensure_output_dir()

    # 生成文件名（如果文件被占用则自动加时间戳）
    date_label = current_date.replace("-", "")
    filename = f"连板股复盘_{date_label}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    # 如果文件存在且被占用，生成带时间戳的新文件名
    if os.path.exists(filepath):
        timestamp = datetime.now().strftime("%H%M%S")
        filename = f"连板股复盘_{date_label}_{timestamp}.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)

    # 创建 DataFrame
    df = pd.DataFrame(stock_list)

    # 自由流通市值 vs 流通市值 回退：自动切换列名
    cap_label = "自由流通盘金额" if is_using_real_free_float() else "流通市值"
    col_rename = {}
    if "自由流通盘金额" in df.columns and cap_label != "自由流通盘金额":
        col_rename["自由流通盘金额"] = cap_label
    ratio_old = "二板竞价成交量/自由流通盘金额"
    ratio_new = f"二板竞价成交量/{cap_label}"
    if ratio_old in df.columns and ratio_old != ratio_new:
        col_rename[ratio_old] = ratio_new
    if col_rename:
        df.rename(columns=col_rename, inplace=True)

    # 构建实际显示列 + 反向映射
    display_columns = [col_rename.get(col, col) for col in COLUMNS]
    col_lookup = {v: k for k, v in col_rename.items()}  # display name → internal key

    # 确保列顺序
    for col in display_columns:
        if col not in df.columns:
            df[col] = "N/A"
    df = df[display_columns]

    # ---- 使用 openpyxl 写入格式化 Excel ----
    wb = Workbook()
    ws = wb.active
    ws.title = "连板股复盘"

    # === 样式定义 ===
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="微软雅黑", size=10)
    data_alignment = Alignment(horizontal="center", vertical="center")
    data_alignment_left = Alignment(horizontal="left", vertical="center")
    data_alignment_right = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # 比值列（百分比/小数格式）
    ratio_cols = {
        "二板竞价涨幅",
        "一板竞价涨幅",
        ratio_new,  # 依赖列名跟着 cap_label 变
        "二板/一板交易量",
        "二板/一板竞价成交量",
        "二板/一板竞价涨幅",
    }

    # 金额列
    amount_cols = {
        cap_label,  # 表头跟着数据源切换
        "二板竞价成交量",
        "二板交易量",
        "一板竞价成交量",
        "一板交易量",
    }

    # === 写入标题行 ===
    title = f"连板股复盘报告 — {prev_date}（一板）→ {current_date}（二板）"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(display_columns))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # 副标题
    subtitle = f"共筛选出 {len(stock_list)} 只连板股 | 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(display_columns))
    sub_cell = ws.cell(row=2, column=1, value=subtitle)
    sub_cell.font = Font(name="微软雅黑", size=9, color="808080")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # === 写入表头（第3行）===
    header_row = 3
    for col_idx, col_name in enumerate(display_columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    ws.row_dimensions[header_row].height = 40

    # === 写入数据（从第4行开始）===
    for row_idx, stock in enumerate(stock_list):
        excel_row = row_idx + 4
        ws.row_dimensions[excel_row].height = 28

        # 交替行颜色
        if row_idx % 2 == 0:
            row_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
        else:
            row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        for col_idx, col_name in enumerate(display_columns, start=1):
            lookup_key = col_lookup.get(col_name, col_name)
            value = stock.get(lookup_key, "N/A")
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.fill = row_fill

            # 对齐方式
            if col_name in ratio_cols:
                cell.alignment = data_alignment_right
                # 百分比格式
                if isinstance(value, (int, float)):
                    if "涨幅" in col_name and "比值" not in col_name and "/" not in col_name:
                        cell.number_format = '0.00"%"'
                    else:
                        cell.number_format = "0.0000"
            elif col_name in amount_cols:
                cell.alignment = data_alignment_right
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0'
            elif col_name == "股票代码":
                cell.alignment = data_alignment
                # 文本格式防止科学计数法
                cell.number_format = '@'
            elif col_name == "股票价格":
                cell.alignment = data_alignment_right
                if isinstance(value, (int, float)):
                    cell.number_format = "0.00"
            elif col_name in ("当天最后涨停时间",):
                cell.alignment = data_alignment
            elif col_name in ("股票名称", "行业"):
                cell.alignment = data_alignment_left
            else:
                cell.alignment = data_alignment

    # === 设置列宽 ===
    col_widths = {
        "股票代码": 12,
        "股票名称": 12,
        "股票价格": 10,
        "行业涨停数": 10,
        "自由流通盘金额": 16,
        "流通市值": 16,
        "行业": 12,
        "当天最后涨停时间": 16,
        "二板竞价涨幅": 14,
        "二板竞价成交量": 16,
        "二板交易量": 16,
        "一板竞价涨幅": 14,
        "一板竞价成交量": 16,
        "一板交易量": 16,
        "二板竞价成交量/自由流通盘金额": 22,
        "二板竞价成交量/流通市值": 22,
        "二板/一板交易量": 16,
        "二板/一板竞价成交量": 18,
        "二板/一板竞价涨幅": 16,
    }

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        width = col_widths.get(col_name, 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # === 冻结表头 ===
    ws.freeze_panes = f"A{header_row + 1}"

    # === 自动筛选 ===
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(COLUMNS))}{len(stock_list) + header_row}"

    # 保存文件
    wb.save(filepath)

    return filepath
